# Location: services/core-backend/app/core/reports/excel_generator.py
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_financial_ledger_excel(
    ledgers_data: list,
    institution_name: str = "Suffat-ul Huffaz",
    report_period: str = "Q3 2026",
) -> bytes:
    """
    Serializes multi-tab double-entry financial ledger data into a
    professionally formatted Excel workbook using openpyxl.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General Ledger"

    # Style definitions
    title_font = Font(name="Segoe UI", size=14, bold=True, color="0D9488")
    subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="64748B")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    data_font = Font(name="Segoe UI", size=10)
    currency_font = Font(name="Segoe UI", size=10, color="1A202C")
    total_font = Font(name="Segoe UI", size=11, bold=True, color="0D9488")
    double_border = Border(
        bottom=Side(style="double", color="1E293B"),
        top=Side(style="thin", color="CBD5E1"),
    )
    thin_border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
    )

    # Title Block
    ws.cell(row=2, column=2, value=f"{institution_name} - Financial Systems").font = title_font
    ws.cell(row=3, column=2, value=f"General Ledger Account Balances — {report_period}").font = subtitle_font

    # Column Headers
    headers = [
        "Account Code",
        "Account Name",
        "Account Type",
        "Debit Balance",
        "Credit Balance",
    ]
    for col_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data Rows
    row_idx = 6
    for entry in ledgers_data:
        ws.cell(row=row_idx, column=2, value=entry.get("code", "")).font = data_font
        ws.cell(row=row_idx, column=3, value=entry.get("name", "")).font = data_font
        ws.cell(row=row_idx, column=4, value=entry.get("type", "")).font = data_font

        debit_cell = ws.cell(row=row_idx, column=5, value=float(entry.get("debit", 0)))
        debit_cell.font = currency_font
        debit_cell.number_format = "$#,##0.00"

        credit_cell = ws.cell(row=row_idx, column=6, value=float(entry.get("credit", 0)))
        credit_cell.font = currency_font
        credit_cell.number_format = "$#,##0.00"

        for c in range(2, 7):
            ws.cell(row=row_idx, column=c).border = thin_border

        row_idx += 1

    # Totals Row
    ws.cell(row=row_idx, column=3, value="TOTAL").font = total_font

    total_debit = ws.cell(row=row_idx, column=5, value=f"=SUM(E6:E{row_idx - 1})")
    total_debit.font = total_font
    total_debit.number_format = "$#,##0.00"
    total_debit.border = double_border

    total_credit = ws.cell(row=row_idx, column=6, value=f"=SUM(F6:F{row_idx - 1})")
    total_credit.font = total_font
    total_credit.number_format = "$#,##0.00"
    total_credit.border = double_border

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Serialize
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes
